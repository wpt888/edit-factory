"""Flexible product tabular import helpers.

The first non-empty row is always treated as the header.  Importers preserve
every source column in ``extra_fields`` while also mapping selected columns to
the canonical fields used by the pipeline.

Imported text is converted from HTML to plain text (see ``html_to_text``).
Every e-commerce platform (GoMag, Shopify, WooCommerce, ...) exports the product
description as an HTML body, and both the canonical ``description`` and the
``extra_fields`` bag are fed verbatim into the AI prompt, so raw markup would
burn tokens and derail the script.
"""

from __future__ import annotations

import csv
import html
import io
import json
import re
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse


CANONICAL_FIELDS = (
    "title",
    "description",
    "external_id",
    "images",
    "brand",
    "category",
    "sku",
    "price",
    "sale_price",
    "product_url",
)

_ALIASES = {
    "title": ("title", "name", "product", "product name", "produs", "nume", "nume produs", "service", "serviciu", "item", "oferta", "ofertă"),
    "description": ("description", "descriere", "body", "details"),
    "external_id": ("external_id", "external id", "id", "product id", "item id"),
    "images": ("images", "image", "image_url", "image url", "image_link", "poze", "imagine"),
    "brand": ("brand", "vendor", "marca"),
    "category": ("category", "product_type", "product type", "categorie"),
    "sku": ("sku", "cod", "product code"),
    "price": ("price", "pret", "preț"),
    "sale_price": ("sale_price", "sale price", "discount price", "pret redus", "preț redus"),
    "product_url": ("product_url", "product url", "link", "url"),
}


# Tags that render as a line break. Everything else is inline and must NOT
# introduce whitespace, or words get split mid-sentence.
_BLOCK_TAGS = frozenset({
    "p", "br", "div", "li", "ul", "ol", "tr", "td", "th", "table",
    "h1", "h2", "h3", "h4", "h5", "h6", "section", "article", "hr", "blockquote",
})

# Deliberately strict: requires a real tag (closing, or opening WITH a '>') or a
# character entity. A bare '<' — "marime < 5", "<XL" — must not look like markup,
# otherwise HTMLParser would swallow it as a bogus tag and silently eat data.
_MARKUP_RE = re.compile(
    r"</[a-zA-Z][^>]*>"
    r"|<[a-zA-Z][a-zA-Z0-9]*(?:\s[^>]*)?/?>"
    r"|&(?:[a-zA-Z]+|#\d+);"
)


class _TextExtractor(HTMLParser):
    """HTML body -> readable plain text; stdlib handles entity decoding."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: Any) -> None:
        if tag in _BLOCK_TAGS:
            self._parts.append("\n")
        if tag == "li":
            self._parts.append("- ")

    def handle_endtag(self, tag: str) -> None:
        if tag in _BLOCK_TAGS:
            self._parts.append("\n")

    def handle_data(self, data: str) -> None:
        self._parts.append(data)

    def text(self) -> str:
        return "".join(self._parts)


def html_to_text(value: str) -> str:
    """Convert an HTML fragment to plain text, preserving line structure.

    A naive ``re.sub(r'<[^>]*>', '', s)`` glues adjacent block text together —
    ``<li>lentila dubla</li><li>lentila transparenta</li>`` collapses into
    "...dublalentila...". Block tags therefore become newlines and list items
    get a "- " bullet. Returns non-markup input untouched (fast path), so plain
    CSV cells (SKUs, prices, sizes like "<XL") are never rewritten. Idempotent.
    """
    if not value or not _MARKUP_RE.search(value):
        return value
    parser = _TextExtractor()
    parser.feed(value)
    parser.close()
    # convert_charrefs already decoded entities; a second unescape catches the
    # double-encoded (&amp;nbsp;) rows common in migrated catalogs. NBSP is then
    # folded to a real space so the collapse below actually sees it.
    out = html.unescape(parser.text()).replace("\xa0", " ")
    out = re.sub(r"[ \t]+", " ", out)
    out = re.sub(r" *\n *", "\n", out)
    out = re.sub(r"\n{3,}", "\n\n", out)
    lines = [line for line in (raw.rstrip() for raw in out.split("\n")) if line not in ("", "-")]
    return "\n".join(lines).strip()


def _clean(value: Any) -> Any:
    """html_to_text for strings; other cell types (int/float/date) pass through."""
    return html_to_text(value) if isinstance(value, str) else value


def google_sheet_csv_url(url: str) -> str:
    """Convert common Google Sheets share URLs to a CSV export URL."""
    parsed = urlparse(url.strip())
    if parsed.netloc.lower() not in {"docs.google.com", "www.docs.google.com"}:
        return url.strip()
    match = re.search(r"/spreadsheets/d/([^/]+)", parsed.path)
    if not match:
        return url.strip()
    query = parse_qs(parsed.query)
    gid = query.get("gid", [None])[0]
    if not gid and parsed.fragment:
        gid = parse_qs(parsed.fragment).get("gid", [None])[0]
    export_query = {"format": "csv"}
    if gid:
        export_query["gid"] = gid
    return urlunparse(("https", "docs.google.com", f"/spreadsheets/d/{match.group(1)}/export", "", urlencode(export_query), ""))


def _clean_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        base = str(value or "").strip() or f"column_{index + 1}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def parse_csv_bytes(data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    raw_rows = list(csv.reader(io.StringIO(text), dialect))
    rows = [row for row in raw_rows if any(str(cell).strip() for cell in row)]
    if not rows:
        return [], []
    headers = _clean_headers(rows[0])
    result = [
        {header: (row[index].strip() if index < len(row) else "") for index, header in enumerate(headers)}
        for row in rows[1:]
    ]
    return headers, result


def parse_xlsx_bytes(data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    values = [list(row) for row in sheet.iter_rows(values_only=True)]
    rows = [row for row in values if any(value not in (None, "") for value in row)]
    if not rows:
        return [], []
    headers = _clean_headers(rows[0])
    result = [
        {header: (row[index] if index < len(row) and row[index] is not None else "") for index, header in enumerate(headers)}
        for row in rows[1:]
    ]
    return headers, result


def parse_google_shopping_xml(data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    from app.services.feed_parser import parse_feed_xml

    rows = parse_feed_xml(data)
    headers = list(dict.fromkeys(key for row in rows for key in row))
    return headers, rows


def parse_product_data(data: bytes, source_type: str, filename: str = "") -> tuple[list[str], list[dict[str, Any]]]:
    normalized = source_type.lower().strip()
    suffix = Path(filename).suffix.lower()
    if normalized == "xlsx" or suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx_bytes(data)
    if normalized in {"xml", "google_shopping_xml"} or suffix == ".xml":
        return parse_google_shopping_xml(data)
    return parse_csv_bytes(data)


def suggest_mapping(headers: list[str]) -> dict[str, Any]:
    normalized = {header.casefold().strip(): header for header in headers}
    mapping: dict[str, Any] = {}
    for field, aliases in _ALIASES.items():
        matches = [normalized[alias.casefold()] for alias in aliases if alias.casefold() in normalized]
        if matches:
            mapping[field] = matches if field == "images" else matches[0]
    return mapping


def _image_values(row: dict[str, Any], columns: Any) -> list[str]:
    if isinstance(columns, str):
        columns = [columns]
    images: list[str] = []
    for column in columns or []:
        raw = str(row.get(column) or "")
        for value in re.split(r"[|,\n]", raw):
            value = value.strip()
            if value and value not in images:
                images.append(value)
    return images


def normalize_rows(rows: list[dict[str, Any]], mapping: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Map rows without discarding source columns; invalid rows are reported."""
    title_column = mapping.get("title")
    if not title_column:
        raise ValueError("Map one source column to Product name")
    products: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        title = html_to_text(str(row.get(title_column) or "")).strip()
        if not title:
            errors.append({"row": row_number, "error": "Product name is empty"})
            continue
        canonical: dict[str, Any] = {
            "title": title,
            "description": html_to_text(str(row.get(mapping.get("description")) or "")).strip(),
            "external_id": str(row.get(mapping.get("external_id")) or row.get(mapping.get("sku")) or "").strip(),
            "brand": str(row.get(mapping.get("brand")) or "").strip(),
            "category": str(row.get(mapping.get("category")) or "").strip(),
            "sku": str(row.get(mapping.get("sku")) or "").strip(),
            "price": str(row.get(mapping.get("price")) or "").strip(),
            "sale_price": str(row.get(mapping.get("sale_price")) or "").strip(),
            "product_url": str(row.get(mapping.get("product_url")) or "").strip(),
            "image_links": _image_values(row, mapping.get("images")),
            # Cleaned too: the prompt builder dumps every non-title/description
            # key verbatim, so a source column named "Descriere" (not matching the
            # builder's exclusion set) would otherwise leak raw HTML into the AI.
            "extra_fields": {key: _clean(value) for key, value in row.items()},
        }
        products.append(canonical)
    return products, errors


def decode_mapping(value: str) -> dict[str, Any]:
    mapping = json.loads(value or "{}")
    if not isinstance(mapping, dict):
        raise ValueError("Mapping must be a JSON object")
    return mapping
